"""Đọc danh sách đơn hàng từ file CSV/Excel và ánh xạ sang field crm.lead."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List

# Ánh xạ cột trong file -> field crm.lead trên Odoo.
# Chỉnh lại cho khớp file thực tế của bạn.
COLUMN_MAP: Dict[str, str] = {
    "ma_don": "name",            # mã đơn -> tên cơ hội / tham chiếu
    "khach_hang": "contact_name",
    "so_dien_thoai": "phone",
    "email": "email_from",
    "doanh_thu": "expected_revenue",
    "ghi_chu": "description",
}

# Cột dùng làm khóa để upsert (match cơ hội đã có).
REF_COLUMN = "ma_don"
REF_FIELD = "name"


def _clean(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
    return value


def load_orders_csv(path: str | Path) -> List[Dict[str, Any]]:
    """Đọc file CSV UTF-8 có dòng tiêu đề."""
    rows: List[Dict[str, Any]] = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        for raw in reader:
            rows.append({k: _clean(v) for k, v in raw.items()})
    return rows


def load_orders_xlsx(path: str | Path) -> List[Dict[str, Any]]:
    """Đọc sheet đầu tiên của file .xlsx (dòng đầu là tiêu đề)."""
    from openpyxl import load_workbook  # import trễ để không bắt buộc cài

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        rows.append({header[i]: _clean(v) for i, v in enumerate(row) if i < len(header)})
    return rows


def load_orders(path: str | Path) -> List[Dict[str, Any]]:
    """Tự nhận định dạng theo đuôi file."""
    suffix = Path(path).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return load_orders_xlsx(path)
    if suffix in (".csv", ".tsv"):
        return load_orders_csv(path)
    raise ValueError(f"Định dạng chưa hỗ trợ: {suffix}")


def order_to_opportunity(order: Dict[str, Any]) -> Dict[str, Any]:
    """Chuyển 1 dòng đơn hàng thành dict field->value của crm.lead."""
    values: Dict[str, Any] = {}
    for col, field in COLUMN_MAP.items():
        if col in order and order[col] not in (None, ""):
            values[field] = order[col]
    # Chuẩn hoá doanh thu về số
    if "expected_revenue" in values:
        try:
            values["expected_revenue"] = float(
                str(values["expected_revenue"]).replace(",", "").replace(" ", "")
            )
        except (TypeError, ValueError):
            values.pop("expected_revenue", None)
    return values
