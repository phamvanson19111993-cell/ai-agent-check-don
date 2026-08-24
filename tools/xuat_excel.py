"""Xuất file Excel 2 sheet: toàn bộ số chưa chốt + riêng page Phúc Thịnh."""

import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(BASE, "data", "sdt_chua_chot_tong_hop.csv")
OUT = os.path.join(BASE, "data", "SDT_chua_chot_Pancake.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WIDTHS = [28, 14, 12, 10, 24, 46]


def write_sheet(sheet, header, rows):
    sheet.append(header)
    for index, cell in enumerate(sheet[1], start=1):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = WIDTHS[index - 1]

    for row in rows:
        sheet.append(row)

    # Cột SĐT để dạng text -> không mất số 0 đầu.
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "@"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:%s%s" % (get_column_letter(len(header)), sheet.max_row)


def main():
    with open(SRC, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        rows = list(reader)

    workbook = Workbook()
    sheet_all = workbook.active
    sheet_all.title = "Tất cả chưa chốt"
    write_sheet(sheet_all, header, rows)

    phuc_thinh = [row for row in rows if row[4].startswith("BPT")]
    write_sheet(workbook.create_sheet("Page Phúc Thịnh"), header, phuc_thinh)

    workbook.save(OUT)
    print("Đã ghi %s" % OUT)
    print("  Sheet 'Tất cả chưa chốt' : %s số" % len(rows))
    print("  Sheet 'Page Phúc Thịnh'  : %s số" % len(phuc_thinh))


if __name__ == "__main__":
    main()
